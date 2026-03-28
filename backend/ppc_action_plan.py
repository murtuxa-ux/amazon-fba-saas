"""
Ecom Era — PPC Action Plan Engine
Automated weekly PPC action plan generation with per-client configurable rules.
Processes Seller Central CSVs / Helium 10 exports / Amazon Ads API data.
Generates: Bid Changes, Keyword Harvests, Negatives, Summary.
AM reviews & edits before applying. Excel export in branded format.
"""

import os, json, io, logging, csv, math
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import Column, Integer, String, Text, Boolean, DateTime, Float, JSON, ForeignKey, desc
from database import get_db, Base, engine
from auth import get_current_user
from models import User

_log = logging.getLogger("ppc_action_plan")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SQLAlchemy Models
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class PPCRulesConfig(Base):
    __tablename__ = "ppc_rules_config"
    __table_args__ = {"extend_existing": True}

    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(Integer, nullable=False, index=True)
    client_id = Column(Integer, nullable=True)
    name = Column(String(200), default="Default Rules")
    target_acos = Column(Float, default=25.0)
    max_acos_hard = Column(Float, default=40.0)
    raise_bid_threshold = Column(Float, default=18.75)
    lower_bid_mild_threshold = Column(Float, default=28.75)
    lower_bid_hard_threshold = Column(Float, default=40.0)
    lower_bid_mild_pct = Column(Float, default=20.0)
    lower_bid_hard_pct = Column(Float, default=35.0)
    raise_bid_pct = Column(Float, default=15.0)
    negative_min_clicks = Column(Integer, default=8)
    negative_min_orders = Column(Integer, default=0)
    harvest_min_orders = Column(Integer, default=1)
    harvest_max_acos = Column(Float, default=35.0)
    min_impressions = Column(Integer, default=100)
    lookback_days = Column(Integer, default=30)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class PPCActionPlan(Base):
    __tablename__ = "ppc_action_plans"
    __table_args__ = {"extend_existing": True}

    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(Integer, nullable=False, index=True)
    client_id = Column(Integer, nullable=True)
    client_name = Column(String(200), nullable=True)
    rules_config_id = Column(Integer, nullable=True)
    plan_date = Column(DateTime, default=datetime.utcnow)
    report_period = Column(String(100), nullable=True)
    status = Column(String(50), default="draft")  # draft, reviewed, approved, applied
    total_keywords = Column(Integer, default=0)
    overall_acos = Column(Float, default=0.0)
    total_spend = Column(Float, default=0.0)
    total_sales = Column(Float, default=0.0)
    bid_changes_count = Column(Integer, default=0)
    harvest_count = Column(Integer, default=0)
    negatives_count = Column(Integer, default=0)
    approved_by = Column(String(200), nullable=True)
    approved_at = Column(DateTime, nullable=True)
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class PPCBidChange(Base):
    __tablename__ = "ppc_bid_changes"
    __table_args__ = {"extend_existing": True}

    id = Column(Integer, primary_key=True, autoincrement=True)
    plan_id = Column(Integer, ForeignKey("ppc_action_plans.id"), nullable=False, index=True)
    campaign = Column(String(300), nullable=True)
    ad_group = Column(String(300), nullable=True)
    keyword_target = Column(String(500), nullable=True)
    match_type = Column(String(50), nullable=True)
    impressions = Column(Integer, default=0)
    clicks = Column(Integer, default=0)
    spend = Column(Float, default=0.0)
    sales = Column(Float, default=0.0)
    orders = Column(Integer, default=0)
    acos = Column(Float, default=0.0)
    current_bid = Column(Float, default=0.0)
    action = Column(String(100), nullable=True)
    new_bid = Column(Float, default=0.0)
    approved = Column(Boolean, nullable=True)
    am_notes = Column(Text, nullable=True)


class PPCKeywordHarvest(Base):
    __tablename__ = "ppc_keyword_harvests"
    __table_args__ = {"extend_existing": True}

    id = Column(Integer, primary_key=True, autoincrement=True)
    plan_id = Column(Integer, ForeignKey("ppc_action_plans.id"), nullable=False, index=True)
    search_term = Column(String(500), nullable=True)
    source_campaign = Column(String(300), nullable=True)
    source_ad_group = Column(String(300), nullable=True)
    auto_target = Column(String(300), nullable=True)
    clicks = Column(Integer, default=0)
    orders = Column(Integer, default=0)
    spend = Column(Float, default=0.0)
    sales = Column(Float, default=0.0)
    recommended_action = Column(String(200), nullable=True)
    match_type = Column(String(50), default="Exact")
    approved = Column(Boolean, nullable=True)
    am_notes = Column(Text, nullable=True)


class PPCNegativeKeyword(Base):
    __tablename__ = "ppc_negative_keywords"
    __table_args__ = {"extend_existing": True}

    id = Column(Integer, primary_key=True, autoincrement=True)
    plan_id = Column(Integer, ForeignKey("ppc_action_plans.id"), nullable=False, index=True)
    search_term = Column(String(500), nullable=True)
    source_campaign = Column(String(300), nullable=True)
    source_ad_group = Column(String(300), nullable=True)
    wasted_clicks = Column(Integer, default=0)
    wasted_spend = Column(Float, default=0.0)
    recommended_action = Column(String(200), nullable=True)
    add_to = Column(String(200), default="All Campaigns")
    priority = Column(String(50), default="High")
    approved = Column(Boolean, nullable=True)
    am_notes = Column(Text, nullable=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Pydantic Schemas
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class RulesConfigInput(BaseModel):
    client_id: Optional[int] = None
    name: str = "Default Rules"
    target_acos: float = 25.0
    max_acos_hard: float = 40.0
    raise_bid_threshold: float = Field(default=18.75, description="ACoS below this = raise bid")
    lower_bid_mild_threshold: float = Field(default=28.75, description="ACoS above this = lower bid mildly")
    lower_bid_hard_threshold: float = Field(default=40.0, description="ACoS above this = lower bid hard")
    lower_bid_mild_pct: float = 20.0
    lower_bid_hard_pct: float = 35.0
    raise_bid_pct: float = 15.0
    negative_min_clicks: int = 8
    negative_min_orders: int = 0
    harvest_min_orders: int = 1
    harvest_max_acos: float = 35.0
    min_impressions: int = 100
    lookback_days: int = 30

class RulesConfigResponse(BaseModel):
    id: int
    client_id: Optional[int]
    name: str
    target_acos: float
    max_acos_hard: float
    raise_bid_threshold: float
    lower_bid_mild_threshold: float
    lower_bid_hard_threshold: float
    lower_bid_mild_pct: float
    lower_bid_hard_pct: float
    raise_bid_pct: float
    negative_min_clicks: int
    harvest_min_orders: int
    harvest_max_acos: float
    min_impressions: int
    lookback_days: int

class ActionPlanSummary(BaseModel):
    id: int
    client_name: Optional[str]
    plan_date: str
    status: str
    overall_acos: float
    total_spend: float
    total_sales: float
    total_keywords: int
    bid_changes_count: int
    harvest_count: int
    negatives_count: int

class BidChangeItem(BaseModel):
    id: int
    campaign: Optional[str]
    ad_group: Optional[str]
    keyword_target: Optional[str]
    match_type: Optional[str]
    impressions: int
    clicks: int
    spend: float
    sales: float
    orders: int
    acos: float
    current_bid: float
    action: Optional[str]
    new_bid: float
    approved: Optional[bool]
    am_notes: Optional[str]

class ApprovalUpdate(BaseModel):
    approved: bool
    am_notes: Optional[str] = None

class GenerateRequest(BaseModel):
    client_id: Optional[int] = None
    client_name: str
    rules_config_id: Optional[int] = None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Rules Engine — Core Logic
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_rules(org_id: int, client_id: int, db: Session) -> PPCRulesConfig:
    rules = None
    if client_id:
        rules = db.query(PPCRulesConfig).filter(
            PPCRulesConfig.org_id == org_id,
            PPCRulesConfig.client_id == client_id,
            PPCRulesConfig.is_active == True
        ).first()
    if not rules:
        rules = db.query(PPCRulesConfig).filter(
            PPCRulesConfig.org_id == org_id,
            PPCRulesConfig.client_id == None,
            PPCRulesConfig.is_active == True
        ).first()
    if not rules:
        rules = PPCRulesConfig(
            org_id=org_id,
            name="Default Rules",
            target_acos=25.0,
            max_acos_hard=40.0,
            raise_bid_threshold=18.75,
            lower_bid_mild_threshold=28.75,
            lower_bid_hard_threshold=40.0,
            lower_bid_mild_pct=20.0,
            lower_bid_hard_pct=35.0,
            raise_bid_pct=15.0,
            negative_min_clicks=8,
            negative_min_orders=0,
            harvest_min_orders=1,
            harvest_max_acos=35.0,
            min_impressions=100,
            lookback_days=30,
        )
        db.add(rules)
        db.commit()
    return rules


def classify_bid_action(acos: float, orders: int, rules: PPCRulesConfig) -> tuple:
    if orders == 0:
        return None, 0.0
    if acos <= 0:
        return None, 0.0
    if acos <= rules.raise_bid_threshold and orders >= 1:
        return "Raise Bid", rules.raise_bid_pct / 100.0
    elif acos > rules.lower_bid_hard_threshold:
        return "Lower Bid (Hard)", -(rules.lower_bid_hard_pct / 100.0)
    elif acos > rules.lower_bid_mild_threshold:
        return "Lower Bid (Mild)", -(rules.lower_bid_mild_pct / 100.0)
    return None, 0.0


def process_keyword_data(rows: List[Dict], rules: PPCRulesConfig) -> Dict[str, List[Dict]]:
    bid_changes = []
    harvests = []
    negatives = []

    total_spend = 0.0
    total_sales = 0.0
    keyword_count = 0

    for row in rows:
        clicks = int(row.get("clicks", 0) or 0)
        impressions = int(row.get("impressions", 0) or 0)
        spend = float(row.get("spend", 0) or 0)
        sales = float(row.get("sales", 0) or row.get("7 Day Total Sales", 0) or 0)
        orders = int(row.get("orders", 0) or row.get("7 Day Total Orders (#)", 0) or 0)
        bid = float(row.get("bid", 0) or row.get("Bid", 0) or 0)
        keyword = row.get("keyword", "") or row.get("Targeting", "") or row.get("Customer Search Term", "") or ""
        campaign = row.get("campaign", "") or row.get("Campaign Name", "") or ""
        ad_group = row.get("ad_group", "") or row.get("Ad Group Name", "") or ""
        match_type = row.get("match_type", "") or row.get("Match Type", "") or ""

        if not keyword.strip():
            continue

        keyword_count += 1
        total_spend += spend
        total_sales += sales
        acos = (spend / sales * 100) if sales > 0 else (999.9 if spend > 0 else 0.0)

        # Bid change analysis
        if impressions >= rules.min_impressions and bid > 0:
            action_label, pct = classify_bid_action(acos, orders, rules)
            if action_label:
                new_bid = round(bid * (1 + pct), 2)
                new_bid = max(0.02, new_bid)
                icon = ""
                if "Raise" in action_label:
                    icon = "Raise Bid"
                elif "Hard" in action_label:
                    icon = "Lower Bid (Hard)"
                else:
                    icon = "Lower Bid (Mild)"
                bid_changes.append({
                    "campaign": campaign,
                    "ad_group": ad_group,
                    "keyword_target": keyword,
                    "match_type": match_type.upper() if match_type else "",
                    "impressions": impressions,
                    "clicks": clicks,
                    "spend": round(spend, 2),
                    "sales": round(sales, 2),
                    "orders": orders,
                    "acos": round(acos, 1),
                    "current_bid": bid,
                    "action": icon,
                    "new_bid": new_bid,
                })

        # Negative keyword analysis (clicks but zero orders)
        if clicks >= rules.negative_min_clicks and orders <= rules.negative_min_orders:
            negatives.append({
                "search_term": keyword,
                "source_campaign": campaign,
                "source_ad_group": ad_group,
                "wasted_clicks": clicks,
                "wasted_spend": round(spend, 2),
                "recommended_action": "Add as NEGATIVE EXACT",
                "add_to": "All Campaigns",
                "priority": "High" if spend > 10 else "Medium",
            })

        # Keyword harvest analysis (converting search terms in auto/broad)
        if orders >= rules.harvest_min_orders and acos <= rules.harvest_max_acos:
            is_auto_or_broad = match_type.lower() in ("broad", "auto", "close", "loose", "substitutes", "complements")
            if is_auto_or_broad:
                harvests.append({
                    "search_term": keyword,
                    "source_campaign": campaign,
                    "source_ad_group": ad_group,
                    "auto_target": match_type,
                    "clicks": clicks,
                    "orders": orders,
                    "spend": round(spend, 2),
                    "sales": round(sales, 2),
                    "recommended_action": "Add as EXACT to Manual Campaign",
                    "match_type": "Exact",
                })

    overall_acos = (total_spend / total_sales * 100) if total_sales > 0 else 0.0

    return {
        "bid_changes": bid_changes,
        "harvests": harvests,
        "negatives": negatives,
        "summary": {
            "total_keywords": keyword_count,
            "overall_acos": round(overall_acos, 1),
            "total_spend": round(total_spend, 2),
            "total_sales": round(total_sales, 2),
        }
    }


def parse_csv_upload(file_bytes: bytes) -> List[Dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        mapped = {}
        for k, v in row.items():
            kl = k.strip().lower()
            if "campaign" in kl:
                mapped["campaign"] = v
            elif "ad group" in kl:
                mapped["ad_group"] = v
            elif "targeting" in kl or "customer search term" in kl or "keyword" in kl:
                mapped["keyword"] = v
            elif "match" in kl and "type" in kl:
                mapped["match_type"] = v
            elif "impression" in kl:
                mapped["impressions"] = v
            elif "click" in kl and "through" not in kl:
                mapped["clicks"] = v
            elif "spend" in kl or "cost" in kl:
                mapped["spend"] = v.replace("$", "").replace(",", "") if v else "0"
            elif "7 day total sales" in kl or "sales" in kl:
                mapped["sales"] = v.replace("$", "").replace(",", "") if v else "0"
            elif "7 day total orders" in kl or "order" in kl:
                mapped["orders"] = v
            elif "bid" in kl:
                mapped["bid"] = v.replace("$", "").replace(",", "") if v else "0"
        rows.append(mapped)
    return rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Excel Export
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_excel_report(plan: PPCActionPlan, bid_changes: list, harvests: list, negatives: list, rules: PPCRulesConfig) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    gold = "FFD700"
    dark_bg = "0A0A0A"
    card_bg = "111111"
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_font = Font(name="Arial", bold=True, size=14, color=gold)
    sub_font = Font(name="Arial", size=10, color="AAAAAA")
    data_font = Font(name="Arial", size=10, color="FFFFFF")
    gold_fill = PatternFill("solid", fgColor=gold)
    dark_fill = PatternFill("solid", fgColor=card_bg)
    header_fill = PatternFill("solid", fgColor="1E1E1E")
    green_font = Font(name="Arial", size=10, color="00FF88")
    red_font = Font(name="Arial", size=10, color="FF4444")
    yellow_font = Font(name="Arial", size=10, color=gold)

    thin_border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = gold

    ws["A1"] = f"ECOM ERA \u2014 PPC ACTION PLAN: {(plan.client_name or 'CLIENT').upper()}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")

    ws["A2"] = f"Generated: {plan.plan_date.strftime('%B %d, %Y')}   |   Report Period: {plan.report_period or 'Last 30 Days'}"
    ws["A2"].font = sub_font

    ws["A4"] = "PERFORMANCE OVERVIEW"
    ws["A4"].font = Font(name="Arial", bold=True, size=12, color=gold)

    summary_data = [
        ("Overall ACoS", f"{plan.overall_acos:.1f}%", f"Target: {rules.target_acos:.0f}%"),
        ("Total Ad Spend", f"${plan.total_spend:,.2f}", ""),
        ("Total Ad Sales", f"${plan.total_sales:,.2f}", ""),
        ("Keywords Analyzed", str(plan.total_keywords), ""),
        ("", "", ""),
        ("ACTION ITEMS", "", ""),
        ("Bid Changes", str(plan.bid_changes_count), ""),
        ("Keywords to Harvest", str(plan.harvest_count), ""),
        ("Negatives to Add", str(plan.negatives_count), ""),
    ]
    for i, (label, val, note) in enumerate(summary_data, 5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
        ws[f"C{i}"] = note
        ws[f"A{i}"].font = data_font if label else Font(name="Arial", size=10)
        ws[f"B{i}"].font = yellow_font if "ACoS" in label else data_font
        ws[f"C{i}"].font = sub_font

    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 25

    # ── Bid Changes Sheet ──
    ws2 = wb.create_sheet("Bid Changes")
    ws2.sheet_properties.tabColor = "FF8800"
    bid_headers = ["Campaign", "Ad Group", "Keyword/Target", "Match", "Impr.", "Clicks", "Spend ($)", "Sales ($)", "Orders", "ACoS", "Current Bid ($)", "Action", "New Bid ($)", "Approved?"]
    ws2["A1"] = f"BID CHANGE RECOMMENDATIONS \u2014 {(plan.client_name or 'CLIENT').upper()}"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:N1")
    ws2["A2"] = "Review each row. Type YES in 'Approved' column to confirm."
    ws2["A2"].font = sub_font

    for col_idx, h in enumerate(bid_headers, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, bc in enumerate(bid_changes, 5):
        vals = [bc.campaign, bc.ad_group, bc.keyword_target, bc.match_type,
                bc.impressions, bc.clicks, bc.spend, bc.sales, bc.orders,
                f"{bc.acos:.1f}%", bc.current_bid, bc.action, bc.new_bid,
                "YES" if bc.approved else ("NO" if bc.approved is False else "")]
        for col_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 12:
                if "Hard" in str(v):
                    cell.font = red_font
                elif "Mild" in str(v):
                    cell.font = yellow_font
                elif "Raise" in str(v):
                    cell.font = green_font

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["L"].width = 20
    ws2.column_dimensions["N"].width = 12

    # ── Keyword Harvest Sheet ──
    ws3 = wb.create_sheet("Keyword Harvest")
    ws3.sheet_properties.tabColor = "00CC66"
    harv_headers = ["Search Term", "Source Campaign", "Source Ad Group", "Auto Target", "Clicks", "Orders", "Spend ($)", "Sales ($)", "Recommended Action", "Match Type", "Approved?"]
    ws3["A1"] = f"KEYWORD HARVEST \u2014 {(plan.client_name or 'CLIENT').upper()}"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:K1")
    ws3["A2"] = "These search terms are converting in Auto/Broad and should be moved to Manual campaigns."
    ws3["A2"].font = sub_font

    for col_idx, h in enumerate(harv_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, kh in enumerate(harvests, 5):
        vals = [kh.search_term, kh.source_campaign, kh.source_ad_group, kh.auto_target,
                kh.clicks, kh.orders, kh.spend, kh.sales,
                kh.recommended_action, kh.match_type,
                "YES" if kh.approved else ("NO" if kh.approved is False else "")]
        for col_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["I"].width = 35

    # ── Negatives Sheet ──
    ws4 = wb.create_sheet("Negatives to Add")
    ws4.sheet_properties.tabColor = "FF4444"
    neg_headers = ["Search Term (to Block)", "Source Campaign", "Source Ad Group", "Wasted Clicks", "Wasted Spend ($)", "Recommended Action", "Add To", "Priority", "Approved?"]
    ws4["A1"] = f"NEGATIVE KEYWORDS \u2014 {(plan.client_name or 'CLIENT').upper()}"
    ws4["A1"].font = title_font
    ws4.merge_cells("A1:I1")
    ws4["A2"] = "These search terms got clicks but ZERO orders. Block them to stop wasting budget."
    ws4["A2"].font = sub_font

    for col_idx, h in enumerate(neg_headers, 1):
        cell = ws4.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, neg in enumerate(negatives, 5):
        vals = [neg.search_term, neg.source_campaign, neg.source_ad_group,
                neg.wasted_clicks, neg.wasted_spend,
                neg.recommended_action, neg.add_to, neg.priority,
                "YES" if neg.approved else ("NO" if neg.approved is False else "")]
        for col_idx, v in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 8 and v == "High":
                cell.font = red_font

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 25
    ws4.column_dimensions["F"].width = 25

    # ── Rules Sheet ──
    ws5 = wb.create_sheet("Rules")
    ws5.sheet_properties.tabColor = "888888"
    ws5["A1"] = "ECOM ERA PPC RULES ENGINE"
    ws5["A1"].font = title_font
    ws5["A2"] = "Rules applied to generate this report. Adjust per client as needed."
    ws5["A2"].font = sub_font

    rules_data = [
        ("TARGET ACoS", f"{rules.target_acos:.0f}%", "Change per client if needed"),
        ("MAX ACoS (hard lower)", f"{rules.max_acos_hard:.0f}%", "Above this -> lower bid hard"),
        ("Raise bid threshold", f"ACoS < {rules.raise_bid_threshold:.2f}%", f"ACoS {rules.raise_bid_threshold/rules.target_acos*100:.0f}% of target with 1+ order"),
        ("Lower bid (mild)", f"ACoS > {rules.lower_bid_mild_threshold:.2f}%", f"ACoS {rules.lower_bid_mild_threshold/rules.target_acos*100:.0f}% of target"),
        ("Lower bid (hard)", f"ACoS > {rules.lower_bid_hard_threshold:.2f}%", f"ACoS {rules.lower_bid_hard_threshold/rules.target_acos*100:.0f}% of target"),
        ("Raise bid amount", f"+{rules.raise_bid_pct:.0f}%", "Applied to current bid"),
        ("Lower mild amount", f"-{rules.lower_bid_mild_pct:.0f}%", "Applied to current bid"),
        ("Lower hard amount", f"-{rules.lower_bid_hard_pct:.0f}%", "Applied to current bid"),
        ("Negative rule", f"{rules.negative_min_clicks}+ clicks, {rules.negative_min_orders} orders", "Add as negative exact"),
        ("Harvest rule", f"{rules.harvest_min_orders}+ orders, ACoS <= {rules.harvest_max_acos:.0f}%", "Move to manual exact"),
        ("Min impressions", str(rules.min_impressions), "Ignore keywords below this"),
    ]

    for i, (label, val, note) in enumerate(rules_data, 4):
        ws5[f"A{i}"] = label
        ws5[f"B{i}"] = val
        ws5[f"C{i}"] = note
        ws5[f"A{i}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws5[f"B{i}"].font = yellow_font
        ws5[f"C{i}"].font = sub_font

    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 30
    ws5.column_dimensions["C"].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# API Router
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

router = APIRouter(prefix="/ppc-action-plan", tags=["ppc-action-plan"])


# ── Rules Config CRUD ──

@router.get("/rules")
def list_rules(
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(PPCRulesConfig).filter(PPCRulesConfig.org_id == user.org_id)
    if client_id is not None:
        q = q.filter(PPCRulesConfig.client_id == client_id)
    configs = q.all()
    return [{
        "id": c.id, "client_id": c.client_id, "name": c.name,
        "target_acos": c.target_acos, "max_acos_hard": c.max_acos_hard,
        "raise_bid_threshold": c.raise_bid_threshold,
        "lower_bid_mild_threshold": c.lower_bid_mild_threshold,
        "lower_bid_hard_threshold": c.lower_bid_hard_threshold,
        "lower_bid_mild_pct": c.lower_bid_mild_pct,
        "lower_bid_hard_pct": c.lower_bid_hard_pct,
        "raise_bid_pct": c.raise_bid_pct,
        "negative_min_clicks": c.negative_min_clicks,
        "harvest_min_orders": c.harvest_min_orders,
        "harvest_max_acos": c.harvest_max_acos,
        "min_impressions": c.min_impressions,
        "lookback_days": c.lookback_days,
    } for c in configs]


@router.post("/rules")
def save_rules(
    data: RulesConfigInput,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    config = PPCRulesConfig(org_id=user.org_id, **data.dict())
    db.add(config)
    db.commit()
    return {"message": "Rules saved", "id": config.id}


@router.put("/rules/{config_id}")
def update_rules(
    config_id: int,
    data: RulesConfigInput,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    config = db.query(PPCRulesConfig).filter(
        PPCRulesConfig.id == config_id,
        PPCRulesConfig.org_id == user.org_id,
    ).first()
    if not config:
        raise HTTPException(status_code=404, detail="Rules config not found")
    for k, v in data.dict().items():
        setattr(config, k, v)
    config.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "Rules updated"}


# ── Action Plans ──

@router.get("/plans")
def list_plans(
    client_id: Optional[int] = None,
    limit: int = Query(default=20, le=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(PPCActionPlan).filter(PPCActionPlan.org_id == user.org_id)
    if client_id:
        q = q.filter(PPCActionPlan.client_id == client_id)
    plans = q.order_by(desc(PPCActionPlan.plan_date)).limit(limit).all()
    return [{
        "id": p.id, "client_name": p.client_name, "plan_date": str(p.plan_date),
        "status": p.status, "overall_acos": p.overall_acos,
        "total_spend": p.total_spend, "total_sales": p.total_sales,
        "total_keywords": p.total_keywords,
        "bid_changes_count": p.bid_changes_count,
        "harvest_count": p.harvest_count,
        "negatives_count": p.negatives_count,
        "approved_by": p.approved_by,
    } for p in plans]


@router.get("/plans/{plan_id}")
def get_plan_detail(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    plan = db.query(PPCActionPlan).filter(
        PPCActionPlan.id == plan_id,
        PPCActionPlan.org_id == user.org_id,
    ).first()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    bids = db.query(PPCBidChange).filter(PPCBidChange.plan_id == plan_id).all()
    harvs = db.query(PPCKeywordHarvest).filter(PPCKeywordHarvest.plan_id == plan_id).all()
    negs = db.query(PPCNegativeKeyword).filter(PPCNegativeKeyword.plan_id == plan_id).all()

    return {
        "plan": {
            "id": plan.id, "client_name": plan.client_name,
            "plan_date": str(plan.plan_date), "status": plan.status,
            "overall_acos": plan.overall_acos, "total_spend": plan.total_spend,
            "total_sales": plan.total_sales, "total_keywords": plan.total_keywords,
            "bid_changes_count": plan.bid_changes_count,
            "harvest_count": plan.harvest_count,
            "negatives_count": plan.negatives_count,
            "notes": plan.notes, "approved_by": plan.approved_by,
            "report_period": plan.report_period,
        },
        "bid_changes": [{
            "id": b.id, "campaign": b.campaign, "ad_group": b.ad_group,
            "keyword_target": b.keyword_target, "match_type": b.match_type,
            "impressions": b.impressions, "clicks": b.clicks,
            "spend": b.spend, "sales": b.sales, "orders": b.orders,
            "acos": b.acos, "current_bid": b.current_bid,
            "action": b.action, "new_bid": b.new_bid,
            "approved": b.approved, "am_notes": b.am_notes,
        } for b in bids],
        "harvests": [{
            "id": h.id, "search_term": h.search_term,
            "source_campaign": h.source_campaign, "source_ad_group": h.source_ad_group,
            "auto_target": h.auto_target, "clicks": h.clicks, "orders": h.orders,
            "spend": h.spend, "sales": h.sales,
            "recommended_action": h.recommended_action,
            "match_type": h.match_type, "approved": h.approved, "am_notes": h.am_notes,
        } for h in harvs],
        "negatives": [{
            "id": n.id, "search_term": n.search_term,
            "source_campaign": n.source_campaign, "source_ad_group": n.source_ad_group,
            "wasted_clicks": n.wasted_clicks, "wasted_spend": n.wasted_spend,
            "recommended_action": n.recommended_action,
            "add_to": n.add_to, "priority": n.priority,
            "approved": n.approved, "am_notes": n.am_notes,
        } for n in negs],
    }


# ── Generate Plan from CSV Upload ──

@router.post("/generate")
async def generate_plan_from_csv(
    file: UploadFile = File(...),
    client_name: str = Form(...),
    client_id: Optional[int] = Form(None),
    rules_config_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    content = await file.read()
    rows = parse_csv_upload(content)

    if not rows:
        raise HTTPException(status_code=400, detail="No valid data found in CSV. Check format.")

    rules = get_rules(user.org_id, client_id, db)
    if rules_config_id:
        specific = db.query(PPCRulesConfig).filter(
            PPCRulesConfig.id == rules_config_id,
            PPCRulesConfig.org_id == user.org_id,
        ).first()
        if specific:
            rules = specific

    result = process_keyword_data(rows, rules)
    summary = result["summary"]

    plan = PPCActionPlan(
        org_id=user.org_id,
        client_id=client_id,
        client_name=client_name,
        rules_config_id=rules.id if rules.id else None,
        plan_date=datetime.utcnow(),
        report_period=f"Last {rules.lookback_days} Days",
        status="draft",
        total_keywords=summary["total_keywords"],
        overall_acos=summary["overall_acos"],
        total_spend=summary["total_spend"],
        total_sales=summary["total_sales"],
        bid_changes_count=len(result["bid_changes"]),
        harvest_count=len(result["harvests"]),
        negatives_count=len(result["negatives"]),
    )
    db.add(plan)
    db.flush()

    for bc in result["bid_changes"]:
        db.add(PPCBidChange(plan_id=plan.id, **bc))
    for h in result["harvests"]:
        db.add(PPCKeywordHarvest(plan_id=plan.id, **h))
    for n in result["negatives"]:
        db.add(PPCNegativeKeyword(plan_id=plan.id, **n))

    db.commit()

    return {
        "message": "Action plan generated",
        "plan_id": plan.id,
        "summary": summary,
        "bid_changes": len(result["bid_changes"]),
        "harvests": len(result["harvests"]),
        "negatives": len(result["negatives"]),
    }


# ── Approve / Update Items ──

@router.put("/bid-changes/{item_id}/approve")
def approve_bid_change(
    item_id: int,
    data: ApprovalUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    item = db.query(PPCBidChange).filter(PPCBidChange.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Bid change not found")
    item.approved = data.approved
    item.am_notes = data.am_notes
    db.commit()
    return {"message": "Updated"}


@router.put("/harvests/{item_id}/approve")
def approve_harvest(
    item_id: int,
    data: ApprovalUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    item = db.query(PPCKeywordHarvest).filter(PPCKeywordHarvest.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Harvest not found")
    item.approved = data.approved
    item.am_notes = data.am_notes
    db.commit()
    return {"message": "Updated"}


@router.put("/negatives/{item_id}/approve")
def approve_negative(
    item_id: int,
    data: ApprovalUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    item = db.query(PPCNegativeKeyword).filter(PPCNegativeKeyword.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Negative not found")
    item.approved = data.approved
    item.am_notes = data.am_notes
    db.commit()
    return {"message": "Updated"}


@router.put("/plans/{plan_id}/approve-all")
def approve_entire_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    plan = db.query(PPCActionPlan).filter(
        PPCActionPlan.id == plan_id,
        PPCActionPlan.org_id == user.org_id,
    ).first()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    plan.status = "approved"
    plan.approved_by = user.name or user.username
    plan.approved_at = datetime.utcnow()
    db.query(PPCBidChange).filter(PPCBidChange.plan_id == plan_id, PPCBidChange.approved == None).update({"approved": True})
    db.query(PPCKeywordHarvest).filter(PPCKeywordHarvest.plan_id == plan_id, PPCKeywordHarvest.approved == None).update({"approved": True})
    db.query(PPCNegativeKeyword).filter(PPCNegativeKeyword.plan_id == plan_id, PPCNegativeKeyword.approved == None).update({"approved": True})
    db.commit()
    return {"message": "Plan approved", "approved_by": plan.approved_by}


@router.put("/plans/{plan_id}/status")
def update_plan_status(
    plan_id: int,
    status: str = Query(..., description="draft, reviewed, approved, applied"),
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    plan = db.query(PPCActionPlan).filter(
        PPCActionPlan.id == plan_id,
        PPCActionPlan.org_id == user.org_id,
    ).first()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    plan.status = status
    if notes:
        plan.notes = notes
    if status == "approved":
        plan.approved_by = user.name or user.username
        plan.approved_at = datetime.utcnow()
    db.commit()
    return {"message": f"Plan status updated to {status}"}


# ── Excel Export ──

@router.get("/plans/{plan_id}/export")
def export_plan_excel(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    plan = db.query(PPCActionPlan).filter(
        PPCActionPlan.id == plan_id,
        PPCActionPlan.org_id == user.org_id,
    ).first()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    bids = db.query(PPCBidChange).filter(PPCBidChange.plan_id == plan_id).all()
    harvs = db.query(PPCKeywordHarvest).filter(PPCKeywordHarvest.plan_id == plan_id).all()
    negs = db.query(PPCNegativeKeyword).filter(PPCNegativeKeyword.plan_id == plan_id).all()

    rules = get_rules(user.org_id, plan.client_id, db)

    excel_bytes = generate_excel_report(plan, bids, harvs, negs, rules)

    safe_name = (plan.client_name or "Client").replace(" ", "_")
    filename = f"PPC_ActionPlan_{safe_name}_{plan.plan_date.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
